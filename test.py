from abc import ABC
from ast import If
from itertools import count
from re import A
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import pandas as pd
data = pd.read_excel('output_files2.csv')
data.head()

wb = load_workbook('output_files2.csv')
ws = wb.active


def graph(colx, coly, qnum):
    a = 0
    b = 0
    c = 0
    d = 0
    for row in range(2, 999):
        for col in range(colx, coly):
            char = get_column_letter(col)
            
            me=(ws[char + str(row)].value)
            if me == 'A':
                a += 1
            elif me == 'B':
                b += 1
            elif me == 'C':
                c += 1
            elif me == 'D':
                d += 1
            
   # print(f'A = {a}')
 #   print(f'B = {b}')
  #  print(f'C = {c}')
   # print(f'D = {d}')

#---------------------------------just inserted------------------------------------------    
    name = ['A', 'B', 'C', 'D']
    score = [a, b, c, d]
    position = [0, 1, 2, 3]

    plt.bar(position, score, width=0.5, color= 'g')
    plt.title(f'Question {qnum}')
    plt.xticks(position, name)
    plt.savefig(f'Question {qnum}.png')
    plt.show()

graph(2, 3, 1)